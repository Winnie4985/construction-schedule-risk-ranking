"""
01_build_features.py — 從 DSLIB 原始 Excel 排程檔萃取任務層級特徵

輸入:
    DSLIB3.2/Excel/*.xlsx           每個工程專案一個活頁簿(Baseline Schedule + 追蹤期分頁)
    DSLIB研究/DSLIB專案盤點表.csv     哪些專案可用(usable)、屬於哪個 Sector

輸出:
    data/features_ext.csv           任務層級特徵表(39 個 f_ 開頭特徵 + 2 個標籤欄)

特徵分組:
    排程結構   f_dur, f_log_dur, f_dur_vs_median, f_start_month, f_rel_pos
    依賴圖論   f_n_pred, f_n_succ, f_n_upstream, f_n_downstream, f_betweenness
    CPM 浮時   f_float, f_is_critical, f_float_ratio
    併行任務   f_concurrent
    成本結構   f_cost, f_log_cost, f_cost_share, f_cost_per_day,
               f_fixed_ratio, f_res_ratio, f_var_ratio, f_has_resource
    銜接關係   f_n_FS, f_n_SS, f_n_FF, f_max_lag, f_sum_lag, f_has_lag
    WBS 階層   f_wbs_depth, f_n_siblings
    資源明細   f_n_resources
    專案層級常數(僅適合跨專案比較,不適合同專案內排序,詳見 reports/methodology_notes.md):
               f_proj_n_tasks, f_proj_med_dur, f_proj_log_cost, f_proj_span,
               f_early_spi, f_early_cpi, f_early_spit, f_proj_n_tp

標籤:
    y_dur_abs   實際工期 > 計畫工期 即算超時(最終模型使用此標籤)
    y_dur_rel   實際工期超過計畫工期 20% 才算超時(僅作敏感度分析用,最終模型未採用)

用法:
    python 01_build_features.py --dslib-root /path/to/DSLIB --out ../data/features_ext.csv
"""
import argparse
import csv
import glob
import math
import os
import re
import warnings

import networkx as nx
import openpyxl
import pandas as pd

from dslib_calendar import (
    add_working_hours,
    find_header,
    parse_duration_hours,
    read_agenda,
)

warnings.filterwarnings("ignore")

# 最終定案僅使用這三類(59 個專案:土木 32、住宅 16、機構建築 11)
SECTOR_TO_CATEGORY = {
    "construction (civil)": "土木",
    "construction (residential building)": "住宅",
    "construction (institutional building)": "機構建築",
}


def _as_int(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def parse_relations(raw):
    """'1FS;2FS8;3SS' -> [(1,'FS',0), (2,'FS',8), (3,'SS',0)]"""
    if raw is None:
        return []
    out = []
    for part in str(raw).split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(\d+)\s*(FS|SS|FF|SF)?\s*([+-]?\d+)?", part, re.I)
        if m and m.group(1):
            out.append((int(m.group(1)), (m.group(2) or "FS").upper(), int(m.group(3) or 0)))
    return out


def tracking_period_sheets(workbook):
    """回傳按追蹤期順序排列的分頁名稱(TP1, TP2, ...)"""
    found = []
    for name in workbook.sheetnames:
        m = re.match(r"^(?:Project Control - )?TP(\d+)$", name)
        if m:
            found.append((int(m.group(1)), name))
    found.sort()
    return [name for _, name in found]


def load_usable_projects(inventory_csv):
    """讀取專案盤點表,回傳 {專案代碼: 類別} (只保留 usable=True 且屬於三個目標類別)"""
    projects = {}
    with open(inventory_csv, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if row["usable"] != "True":
                continue
            category = SECTOR_TO_CATEGORY.get(row["Sector"].strip().lower())
            if category:
                projects[row["code"]] = category
    return projects


def read_baseline_tasks(worksheet, header_row, col_index, hours_per_day, slots, workdays, holidays):
    """讀取 Baseline Schedule 分頁,回傳 {task_id: 任務屬性 dict}"""
    tasks = {}
    for r in range(header_row + 1, worksheet.max_row + 1):
        task_id = _as_int(worksheet.cell(row=r, column=col_index["ID"]).value)
        if task_id is None:
            continue

        def cell(col_name):
            return worksheet.cell(row=r, column=col_index[col_name]).value if col_name in col_index else None

        start = cell("Baseline Start")
        end = cell("Baseline End")
        duration_h = parse_duration_hours(cell("Duration"), hours_per_day)
        if end is None and hasattr(start, "year") and duration_h is not None:
            end = add_working_hours(start, duration_h, slots, workdays, holidays)

        def as_num(v):
            return float(v) if isinstance(v, (int, float)) else 0.0

        total_cost = cell("Total Cost")
        total_cost = as_num(total_cost) if "Total Cost" in col_index else (
            as_num(cell("Resource Cost")) + as_num(cell("Fixed Cost")) + as_num(cell("Variable Cost"))
        )
        tasks[task_id] = dict(
            wbs=str(cell("WBS") or ""),
            preds=parse_relations(cell("Predecessors")),
            succs=parse_relations(cell("Successors")),
            resource_demand=cell("Resource Demand"),
            start=start if hasattr(start, "year") else None,
            end=end if hasattr(end, "year") else None,
            duration_h=duration_h,
            resource_cost=as_num(cell("Resource Cost")),
            fixed_cost=as_num(cell("Fixed Cost")),
            variable_cost=as_num(cell("Variable Cost")),
            total_cost=total_cost,
        )
    return tasks


def build_dependency_graph(tasks):
    """依 Predecessors 建有向圖,若有環(資料品質問題)則移除環邊以確保 DAG"""
    graph = nx.DiGraph()
    graph.add_nodes_from(tasks)
    for task_id, task in tasks.items():
        for pred_id, _rel, _lag in task["preds"]:
            if pred_id in tasks and pred_id != task_id:
                graph.add_edge(pred_id, task_id)
    if not nx.is_directed_acyclic_graph(graph):
        try:
            while True:
                cycle_edge = nx.find_cycle(graph, orientation="original")
                graph.remove_edge(cycle_edge[0][0], cycle_edge[0][1])
        except nx.NetworkXNoCycle:
            pass
    return graph


def compute_cpm(graph, duration_days):
    """前推(ES/EF)+ 後推(LS/LF)計算浮時,回傳 (float_by_task, project_end)"""
    try:
        order = list(nx.topological_sort(graph))
    except nx.NetworkXUnfeasible:
        order = list(graph.nodes)

    earliest_start, earliest_finish = {}, {}
    for t in order:
        preds = list(graph.predecessors(t))
        earliest_start[t] = max((earliest_finish[p] for p in preds if p in earliest_finish), default=0.0)
        earliest_finish[t] = earliest_start[t] + duration_days.get(t, 0.0)
    project_end = max(earliest_finish.values(), default=0.0)

    latest_finish, latest_start = {}, {}
    for t in reversed(order):
        succs = list(graph.successors(t))
        latest_finish[t] = min((latest_start[s] for s in succs if s in latest_start), default=project_end)
        latest_start[t] = latest_finish[t] - duration_days.get(t, 0.0)

    task_float = {t: latest_start.get(t, 0.0) - earliest_start.get(t, 0.0) for t in duration_days}
    return task_float, project_end


def read_early_tracking_performance(workbook):
    """從 Tracking Overview 分頁抓第一個追蹤期的 SPI/CPI/SPI(t)(專案層級,僅供跨專案比較用)"""
    result = {}
    if "Tracking Overview" not in workbook.sheetnames:
        return result
    sheet = workbook["Tracking Overview"]
    header, header_row = None, None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        values = [str(x) for x in row if x is not None]
        if "Planned Value (PV)" in values or "Earned Value (EV)" in values:
            header = {str(c): j for j, c in enumerate(row) if c is not None}
            header_row = i
            break
    if not header:
        return result
    data_rows = [
        row for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, values_only=True)
        if any(x is not None for x in row)
    ]
    if not data_rows:
        return result
    first = data_rows[0]

    def get(key):
        return first[header[key]] if key in header and header[key] < len(first) else None

    for key, name in [
        ("Schedule Performance Index (SPI)", "spi"),
        ("Cost Performance Index (CPI)", "cpi"),
        ("Schedule Performance Index (SPI(t))", "spit"),
    ]:
        value = get(key)
        result[name] = float(value) if isinstance(value, (int, float)) else float("nan")
    result["n_tp"] = len(data_rows)
    return result


def process_workbook(path, code, category):
    """處理單一專案 Excel 檔,回傳該專案已完工任務的特徵列表(list of dict)"""
    workbook = openpyxl.load_workbook(path, data_only=True)
    agenda = read_agenda(workbook)
    if agenda is None:
        workbook.close()
        return []
    slots, workdays, holidays = agenda
    hours_per_day = len(slots)
    if hours_per_day == 0 or not workdays:
        workbook.close()
        return []

    sheet = workbook["Baseline Schedule"]
    header_row, col_index = find_header(sheet)
    if header_row is None:
        workbook.close()
        return []

    tasks = read_baseline_tasks(sheet, header_row, col_index, hours_per_day, slots, workdays, holidays)
    if len(tasks) < 20:
        workbook.close()
        return []

    graph = build_dependency_graph(tasks)
    try:
        betweenness = nx.betweenness_centrality(graph)
    except Exception:
        betweenness = {n: 0.0 for n in graph}

    duration_days = {
        t: (tasks[t]["duration_h"] / hours_per_day if tasks[t]["duration_h"] is not None else 0.0)
        for t in tasks
    }
    task_float, project_end = compute_cpm(graph, duration_days)

    windows = [(t, tasks[t]["start"], tasks[t]["end"]) for t in tasks if tasks[t]["start"] and tasks[t]["end"]]

    def concurrent_count(task_id):
        task = tasks[task_id]
        if not task["start"] or not task["end"]:
            return 0
        return sum(
            1 for (other_id, s, e) in windows
            if other_id != task_id and s < task["end"] and e > task["start"]
        )

    wbs_depth = {t: (tasks[t]["wbs"].count(".") + 1 if tasks[t]["wbs"] else 0) for t in tasks}
    parent_wbs = {t: (".".join(tasks[t]["wbs"].split(".")[:-1]) if "." in tasks[t]["wbs"] else "") for t in tasks}
    siblings = pd.Series(list(parent_wbs.values())).value_counts().to_dict()

    positive_durations = [d for d in duration_days.values() if d > 0]
    median_duration = sorted(positive_durations)[len(positive_durations) // 2] if positive_durations else 1.0
    total_cost = sum(tasks[t]["total_cost"] for t in tasks) or 1.0
    starts = [tasks[t]["start"] for t in tasks if tasks[t]["start"]]
    ends = [tasks[t]["end"] for t in tasks if tasks[t]["end"]]
    project_start = min(starts) if starts else None
    project_finish = max(ends) if ends else None
    project_span = max((project_finish - project_start).days, 1) if (project_start and project_finish) else 1

    early_perf = read_early_tracking_performance(workbook)

    tp_sheets = tracking_period_sheets(workbook)
    if not tp_sheets:
        workbook.close()
        return []
    last_tp = workbook[tp_sheets[-1]]
    tp_header_row, tp_col_index = find_header(last_tp, must=("ID", "Actual Start"))
    if tp_header_row is None:
        workbook.close()
        return []

    tracking = {}
    for r in range(tp_header_row + 1, last_tp.max_row + 1):
        task_id = _as_int(last_tp.cell(row=r, column=tp_col_index["ID"]).value)
        if task_id is None:
            continue

        def tp_cell(col_name):
            return last_tp.cell(row=r, column=tp_col_index[col_name]).value if col_name in tp_col_index else None

        tracking[task_id] = dict(
            actual_duration_h=parse_duration_hours(tp_cell("Actual Duration"), hours_per_day),
            pct_complete=tp_cell("Percentage Completed"),
        )

    rows = []
    for task_id, task in tasks.items():
        tk = tracking.get(task_id)
        if tk is None:
            continue
        pct = tk["pct_complete"]
        if not isinstance(pct, (int, float)) or pct < 0.999:
            continue  # 只用已完工任務
        planned_days = duration_days.get(task_id, 0.0)
        actual_days = tk["actual_duration_h"] / hours_per_day if tk["actual_duration_h"] is not None else None
        if planned_days <= 0 or actual_days is None:
            continue

        preds, succs = task["preds"], task["succs"]
        lags = [lag for _, _, lag in preds] + [lag for _, _, lag in succs]
        rel_types = [rel for _, rel, _ in preds]
        resource_demand = str(task["resource_demand"] or "")
        n_resources = len([x for x in resource_demand.split(";") if x.strip()]) if resource_demand else 0
        cost = task["total_cost"]

        dur_ratio = (actual_days - planned_days) / planned_days
        rows.append(dict(
            code=code, cat=category, tid=task_id,
            f_dur=planned_days, f_log_dur=math.log1p(planned_days),
            f_dur_vs_median=planned_days / median_duration if median_duration > 0 else 0,
            f_start_month=task["start"].month if task["start"] else 0,
            f_rel_pos=((task["start"] - project_start).days / project_span) if (task["start"] and project_start) else 0,
            f_n_pred=len(list(graph.predecessors(task_id))), f_n_succ=len(list(graph.successors(task_id))),
            f_n_upstream=len(nx.ancestors(graph, task_id)) if task_id in graph else 0,
            f_n_downstream=len(nx.descendants(graph, task_id)) if task_id in graph else 0,
            f_betweenness=betweenness.get(task_id, 0.0),
            f_float=task_float.get(task_id, 0.0),
            f_is_critical=int(abs(task_float.get(task_id, 0.0)) < 1e-6),
            f_float_ratio=task_float.get(task_id, 0.0) / project_end if project_end > 0 else 0,
            f_concurrent=concurrent_count(task_id),
            f_cost=cost, f_log_cost=math.log1p(max(cost, 0)), f_cost_share=cost / total_cost,
            f_cost_per_day=cost / planned_days if planned_days else 0,
            f_fixed_ratio=task["fixed_cost"] / cost if cost else 0,
            f_res_ratio=task["resource_cost"] / cost if cost else 0,
            f_var_ratio=task["variable_cost"] / cost if cost else 0,
            f_has_resource=int(bool(task["resource_demand"])),
            f_n_FS=rel_types.count("FS"), f_n_SS=rel_types.count("SS"), f_n_FF=rel_types.count("FF"),
            f_max_lag=max(lags) if lags else 0, f_sum_lag=sum(lags) if lags else 0,
            f_has_lag=int(any(lag != 0 for lag in lags)),
            f_wbs_depth=wbs_depth.get(task_id, 0),
            f_n_siblings=siblings.get(parent_wbs.get(task_id, ""), 0),
            f_n_resources=n_resources,
            # 專案層級常數(見 methodology_notes.md 的「常數特徵污染」討論)
            f_proj_n_tasks=len(tasks), f_proj_med_dur=median_duration,
            f_proj_log_cost=math.log1p(max(total_cost, 0)), f_proj_span=project_span,
            f_early_spi=early_perf.get("spi", float("nan")), f_early_cpi=early_perf.get("cpi", float("nan")),
            f_early_spit=early_perf.get("spit", float("nan")), f_proj_n_tp=early_perf.get("n_tp", float("nan")),
            y_dur_abs=int(actual_days > planned_days),
            y_dur_rel=int(dur_ratio > 0.20),
        ))
    workbook.close()
    return rows


def run(dslib_root, out_path):
    excel_dir = os.path.join(dslib_root, "DSLIB3.2", "Excel")
    inventory_csv = os.path.join(dslib_root, "DSLIB研究", "DSLIB專案盤點表.csv")

    projects = load_usable_projects(inventory_csv)
    print(f"目標專案數: {len(projects)}")

    all_rows = []
    for path in sorted(glob.glob(os.path.join(excel_dir, "*.xlsx"))):
        code = os.path.basename(path).split(" ")[0].replace(".xlsx", "")
        if code not in projects:
            continue
        all_rows.extend(process_workbook(path, code, projects[code]))

    df = pd.DataFrame(all_rows)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    df.to_csv(out_path, index=False)
    print(f"任務筆數: {len(df)}  專案數: {df['code'].nunique()}")
    print(df["cat"].value_counts())
    print(f"已輸出: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dslib-root", required=True, help="DSLIB 資料集根目錄(內含 DSLIB3.2/、DSLIB研究/)")
    parser.add_argument("--out", default="../data/features_ext.csv", help="輸出特徵表路徑")
    args = parser.parse_args()
    run(args.dslib_root, args.out)
