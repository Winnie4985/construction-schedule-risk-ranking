"""DSLIB 行事曆感知的工期推算工具"""
import openpyxl, re, warnings
from datetime import datetime, timedelta
warnings.filterwarnings('ignore')

DAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

def read_agenda(wb):
    """回傳 (working_hour_slots:set[int], working_weekdays:set[int 0=Mon], holidays:set[date])"""
    if 'Agenda' not in wb.sheetnames:
        return None
    ws = wb['Agenda']
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    slots, wdays, hols = set(), set(), set()
    hour_i = 0
    for r in rows:
        r = list(r) + [None]*(8-len(r))
        # working hours: col A label "H:00 - H:00", col B Yes/No
        if isinstance(r[0], str) and re.match(r'^\d+:00 - \d+:00$', r[0]):
            h = int(r[0].split(':')[0])
            if r[1] == 'Yes':
                slots.add(h)
        if r[3] in DAYS and r[4] == 'Yes':
            wdays.add(DAYS.index(r[3]))
        if r[6] is not None and hasattr(r[6], 'year'):
            hols.add(r[6].date())
    return slots, wdays, hols

def parse_duration_hours(s, hours_per_day):
    """把 '5d' / '3h' / '1d 2h' / 0 轉成小時數"""
    if s is None: return None
    if isinstance(s, (int, float)): 
        return float(s) * hours_per_day if s != 0 else 0.0
    s = str(s).strip()
    if s in ('', '0'): return 0.0
    tot = 0.0
    found = False
    for m in re.finditer(r'(\d+(?:[.,]\d+)?)\s*([dh])', s):
        v = float(m.group(1).replace(',', '.'))
        tot += v * hours_per_day if m.group(2) == 'd' else v
        found = True
    return tot if found else None

def is_working_day(d, wdays, hols):
    return d.weekday() in wdays and d not in hols

def add_working_hours(start, hours, slots, wdays, hols, max_days=20000):
    """從 start 開始消耗 hours 個工作小時,回傳結束時間"""
    if hours is None: return None
    if hours <= 0: return start
    cur = start
    remaining = hours
    guard = 0
    while remaining > 0:
        guard += 1
        if guard > max_days * 24: return None
        d = cur.date()
        if not is_working_day(d, wdays, hols):
            cur = datetime.combine(d + timedelta(days=1), datetime.min.time())
            continue
        h = cur.hour
        if h not in slots:
            nxt = cur.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            cur = nxt
            continue
        # consume this hour (or part)
        frac_used = cur.minute/60.0
        avail = 1.0 - frac_used
        if remaining >= avail:
            remaining -= avail
            cur = cur.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        else:
            cur = cur + timedelta(hours=remaining)
            remaining = 0
    return cur

def working_days_between(a, b, wdays, hols):
    """a 到 b 之間的工作日數(含頭不含尾的日數差,以日為單位)"""
    if a is None or b is None: return None
    sign = 1
    if b < a:
        a, b = b, a
        sign = -1
    n = 0
    d = a.date()
    end = b.date()
    guard = 0
    while d < end:
        guard += 1
        if guard > 40000: return None
        d = d + timedelta(days=1)
        if is_working_day(d, wdays, hols):
            n += 1
    return sign * n

def find_header(ws, must=('ID','Duration'), max_scan=15):
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for i, r in enumerate(rows, 1):
        vals = [str(c) for c in r if c is not None]
        if all(m in vals for m in must):
            hdr = [str(c) if c is not None else None for c in r]
            return i, {h: j+1 for j, h in enumerate(hdr) if h}
    return None, None
